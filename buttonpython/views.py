import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
import requests
import json
import xlwt
import csv
import openpyxl
from django.shortcuts import render
import requests
import xlwt
import csv
import openpyxl
from openpyxl import Workbook


def button(request):
    return render(request,'home.html')


def output(request):
    url_oc = "https://www.nseindia.com/api/equity-stock"
    url = f"https://www.nseindia.com/api/equity-stock?index=allcontracts"
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, '
                             'like Gecko) '
                             'Chrome/80.0.3987.149 Safari/537.36',
               'accept-language': 'en,gu;q=0.9,hi;q=0.8', 'accept-encoding': 'gzip, deflate, br'}
    session = requests.Session()

    request1 = session.get(url_oc, headers=headers, timeout=5)
    cookies = dict(request1.cookies)
    response = session.get(url, headers=headers, timeout=5, cookies=cookies)
    print(type(response.json()))
    test = "Hello"
    #excel code
    data = json.dump(response.json(), open('j2init.json', "w"))
    with open("j2init.json") as jfile:
        jd = json.load(jfile)
        wb = Workbook()
        ws_01 = wb.active
        ws_01.title = "Equity"

        ws_01.cell(1, 1, "InstrumentType")
        ws_01.cell(1, 2, "Symbol")
        ws_01.cell(1, 3, "ExpiryDate")
        ws_01.cell(1, 4, "OptionType")
        ws_01.cell(1, 5, "StrikePrice")
        ws_01.cell(1, 6, "LTP")
        ws_01.cell(1, 7, "%CHNG")
        ws_01.cell(1, 8, "Volume")
        ws_01.cell(1, 9, "value")
        ws_01.cell(1, 10, "OpenInterest")
        ws_01.cell(1, 11, "ValueOfUnderlying")
        ws_01.cell(1, 12, "LastPrice")
        ws_01.cell(1, 13, "HighPrice")
        ws_01.cell(1, 14, "Percentage")

        row = 1
        for item in jd["volume"]:
            row = row + 1
            ws_01.cell(row, 1, item['instrument'])
            ws_01.cell(row, 2, item['underlying'])
            ws_01.cell(row, 3, item['expiryDate'])
            ws_01.cell(row, 4, item['optionType'])
            ws_01.cell(row, 5, item['strikePrice'])
            ws_01.cell(row, 6, item['lastPrice'])
            ws_01.cell(row, 7, item['pChange'])
            ws_01.cell(row, 8, item['numberOfContractsTraded'])
            ws_01.cell(row, 9, item['totalTurnover'])
            ws_01.cell(row, 10, item['openInterest'])
            ws_01.cell(row, 11, item['underlyingValue'])
            ws_01.cell(row, 12, item['lastPrice'])
            ws_01.cell(row, 13, item['highPrice'])
            per = ((item['highPrice'] - item['lastPrice']) / item['highPrice']) * 100
            print(per)
            ws_01.cell(row, 14, per)
            return render(request, 'home.html', {'data': per})
    #till here exccel code

    #data = requests.get("https://www.nseindia.com/api/equity-stock?index=allcontracts")
    #data = data.text
    #data = print("Hello,Its working")
    #print("Hello")
    #return render(request,'home.html',{'data':data})
    #return render(request, 'home.html', {'data': response.json()})

    #return render(request,'home.html')
